# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import logging
import openpyxl
import io
import re
from datetime import datetime, date

from odoo import api, fields, models
from odoo.fields import Command
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)
DATE_FORMAT = "%Y-%m-%d"

def get_date_formated(date):

    if date and re.match(r'\d{4}-\d{1,2}-\d{1,2}', date):
        return date
    elif date and type(date) in [pd.Timestamp]:
       return _date.strftime(DATE_FORMAT)
    elif date and type(date) in [str, datetime, date]:
       _date = datetime.strptime(date, "%d/%m/%Y") if type(date) not in [datetime, date] else date
       return _date.strftime(DATE_FORMAT)
    else:
       return ""


class WorkflowRules(models.Model):
    _inherit = "documents.workflow.rule"

    topnet_action = fields.Selection(
        selection=[
            ('load_emp', "Charger Employe"),
            ('load_sal', 'Charger Salaire'),
            ('load_hol', 'Charger Conge'),
            ('load_ret', 'Charger Retraite Comp.'),
        ],
        string="Action Speciales Topnett",
        default="",
        required=False
    )

    def get_display_notif(self, title, message):
        action= {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': title, 'message': message, 'sticky': True, #True/False will display for few seconds if false
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }
        return action

    def apply_actions(self, document_ids):
        """
        called by the front-end Document Inspector to apply the actions to the selection of ID's.

        :param document_ids: the list of documents to apply the action.
        :return: if the action was to create a new business object, returns an action to open the view of the
                newly created object, else returns True.
        """
        topnett_rules = self.filtered(lambda x: x.topnet_action)
        others = self - topnett_rules
        if self.topnet_action:

            documents = self.env['documents.document'].browse(document_ids)

            # partner/owner/share_link/folder changes
            document_dict = {}
            if self.user_id:
                document_dict['owner_id'] = self.user_id.id
            if self.partner_id:
                document_dict['partner_id'] = self.partner_id.id
            if self.folder_id:
                document_dict['folder_id'] = self.folder_id.id

            # Use sudo if user has write access on document else allow to do the
            # other workflow actions(like: schedule activity, send mail etc...)
            try:
                documents.check_access_rights('write')
                documents.check_access_rule('write')
                documents = documents.sudo()
            except AccessError:
                pass

            documents.write(document_dict)

            msg = ""
            for document in documents:
                if self.remove_activities:
                    document.activity_ids.action_feedback(
                        feedback="completed by rule: %s. %s" % (self.name, self.note or '')
                    )

                if self.topnet_action == "load_emp":
                    msg = self.execute_load_employee(document)
                if self.topnet_action == "load_sal":
                    msg = self.with_delay().execute_load_salaire(document)
                    msg = f"Verifiez si la tache {msg} est terminee"
                if self.topnet_action == "load_hol":
                    msg = self.execute_load_holiday(document)
                if self.topnet_action == "load_ret":
                    msg = self.execute_load_retraite(document)

                # tag and facet actions
                for tag_action in self.tag_action_ids:
                    tag_action.execute_tag_action(document)

            if self.create_model:
                return self.with_company(documents.company_id).create_record(documents=documents)

            return {'message': msg}
        else:
            return super(WorkflowRules, others).apply_actions(document_ids)

    def get_period_importation(self, url_file, filename=""):
        datas = []
        default_period = "Période : 14/23"
        input_b = io.BytesIO(url_file)
        dataframe = openpyxl.load_workbook(filename=input_b)
        dataframe1 = dataframe.active

        month_year = []
        date_salaire = date.today()

        if filename:
            matches = re.findall(r"(\d{2})_(\d{2})", filename)
            if matches:
                month_year = list(map(lambda x: x.isdecimal() and int(x) or 0, matches[0]))
                if len(month_year) < 2 or month_year[0] > 12:
                    month_year = []

        for row in range(0, dataframe1.max_row):
            if len(month_year) > 0:
                    break

            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                if len(month_year) > 0:
                    break

                value = str(col[row].value)
                if value:
                    matches = re.findall(r"[Period|Période] : (\d{2})/(\d{2})", value)
                    if len(matches) > 0:
                        try:
                            month_year = list(map(lambda x: x.isdecimal() and int(x) or 0, matches[0]))
                        except:
                            continue

        if len(month_year) == 2:
            month, year = month_year
            if month > 0 and year > 0:
                date_salaire = date(2000+year, month, 1)
        return date_salaire.strftime(DATE_FORMAT)

    def file_to_dict(self, url, line_header=3, line_first_row=4, handler=None, column_header=[], column_index=[], filename=""):
        """Get the data from URL."""
        _logger.info("Start get excel file ..............................")

        date_salaire = self.get_period_importation(url, filename=filename)
        if handler == 'spreadsheet':
            raise UserError("EXTENSION NON PRISE EN COMPTE")
            excel_file = pd.read_json(url)
        elif handler == 'openxl':
            datas = []
            input_b = io.BytesIO(url)
            dataframe = openpyxl.load_workbook(filename=input_b)
            dataframe1 = dataframe.active
            for row in range(0, dataframe1.max_row):
                line = []
                for col in dataframe1.iter_cols(1, dataframe1.max_column):
                    line.append(col[row].value)
                datas.append(line)
            clean_datas = datas[6:] if len(datas) > 6 else datas
            clean_datas = [line for line in clean_datas if any(line)]
            arr_np = np.array(clean_datas)
            transposed_array = np.transpose(arr_np)
            array_list = transposed_array.tolist()
            array_list = [line for line in array_list if any(line)]
            header = array_list[0]
            datas_traites = array_list[1:]

            zip_datas = []
            i = 0
            if len(datas_traites) % 2 != 0:
                raise UserError("Mauvais fichier les donnees des employes doivent etre en double")

            while i < len(datas_traites):
               new_list = list(zip(datas_traites[i], datas_traites[i+1]))
               zip_datas.append(new_list)
               i += 2

            _clean_zip_datas = [list(map(lambda x: type(x) == tuple and tuple(filter(lambda a: a, x)) or [] ,line)) for line in zip_datas]
            clean_zip_datas = [list(map(lambda x: x and type(x) == tuple and x[-1] or '' ,line)) for line in _clean_zip_datas]

            header[0] = 'matricule'
            header[1] = 'nom'
            header[2] = 'prenom'
            df_data = pd.DataFrame.from_records(clean_zip_datas, columns=header)
            datas = df_data.to_dict('records')
            _logger.debug(f" template line: {str(datas[:0])}")
            _logger.debug("end extract..............................")
            return datas, date_salaire
        else:
            input_b = io.BytesIO(url)
            excel_file = pd.read_excel(input_b, keep_default_na=False, na_values=['nan', 'NAN', 'NaN', 'Nan', 'NA']) # get excel file for salary
            excel_file.replace(np.nan, "")
            excel_file.fillna("")

        _logger.info(excel_file)
        if column_header and column_index:
            excel_file = excel_file.iloc[line_first_row:,column_index]
            excel_file.columns = column_header
            _logger.info(excel_file)
            _datas = excel_file
            datas = _datas.to_dict('records')
            _logger.info(datas)
        else:
            excel_file.columns = excel_file.iloc[line_header]
            _datas = excel_file.iloc[line_first_row:].dropna(axis=1, how='all')
            datas = _datas.to_dict('records')
        
        datas = ['nan' in dt.keys() and dt.pop('nan') and dt or dt for dt in datas]

        _logger.info(f" template line: {str(datas[:0])}")
        _logger.info("end extract..............................")
        return datas, date_salaire

    def transform_employee(self, data_employee):
        """Changement du nom des colonnes employes et adaptation des valeurs"""
        _logger.info("Start transform employee data ..............................")

        def v_get_value(value, key):
            if not value:                
                return value
            elif key == 'numero':
                try:
                    value = value and int(value)
                except ValueError:
                    value = None
                return value
            elif key == 'birthday':                
                return get_date_formated(value)
            elif key == 'registration_number':
                vals = value.split("/") if value and type(value) == str else []
                return len(vals) > 1 and vals[-1].strip() or value
            else:
                return value

        _logger.info("Start transform..............................")
        map_croisement = {'Matricule': 'registration_number', 'Nom': 'nom', 'Prenom': 'prenom', 'Date naissance': 'birthday', 'Numero': 'numero'}
        transformed = []
        for line in data_employee:
            transformed.append({ map_croisement.get(key): v_get_value(line[key], map_croisement.get(key)) for key in line.keys() if key in map_croisement.keys()})

        _logger.info(f"{str(transformed[:5])}")
        _logger.info("end transform..............................")
        return transformed

    def update_employee(self, transformed):
        """Recupere les employee existants et les met a jour."""
        _logger.info("Start update employee ..............................")
        emp_lines = self.sudo().env["hr.employee"].search_read([], ['id', 'numero'])

        existing_products = []
        new_products = []
        for lp_ in transformed:
            if 'numero' not in lp_.keys() or not lp_['numero']:
                continue

            exist = list(filter(lambda x: x['numero'] and x['numero'] == lp_['numero'], emp_lines))
            if exist:
                existing_products.append((exist[0]['id'], lp_))
            else:
                new_products.append(lp_)

        msg = f"{len(existing_products)} employe(s) mis a jour et {len(new_products)} nouveaux employes"

        # write ou update
        for _id, vals in existing_products:
            self.env["hr.employee"].browse(_id).update(vals)

        # create new employee
        for elt in new_products:
            self.env["hr.employee"].create(elt)

        _logger.info("end updating employee..............................")
        return msg

    def transform_holiday(self, data_holiday):
        """Changement du nom des colonnes conges et adaptation des valeurs"""
        _logger.info("Start transform conge data ..............................")

        def v_get_value(value, key):
            if not value:                
                return value
            elif key in ['nb_heures', 'nb_jours']:
                try:
                    value = value and int(value)
                except ValueError:
                    value = None
                return value
            elif key in ['date_start', 'date_end']:                
                return get_date_formated(value)
            elif key == 'registration_number':
                vals = value.split("/") if value and type(value) == str else []
                return len(vals) > 1 and vals[-1].strip() or value
            else:
                return value.strip() if type(value) == str else value

        _logger.info("Start transform..............................")
        map_croisement = {'Motif': 'code_conge', 'Employé': 'matricule', 'Du': 'date_start', 'Au': 'date_end', 'Nb Heures': 'nb_heures', 'Nb Jours': 'nb_jours'}
        transformed = []
        for line in data_holiday:
            transformed.append({ map_croisement.get(key): v_get_value(line[key], map_croisement.get(key)) for key in line.keys() if key in map_croisement.keys()})

        _logger.info(f"{str(transformed[:5])}")
        _logger.info("end transform..............................")
        return transformed

    def update_holiday(self, transformed):
        """Recupere les holidays existants et les met a jour."""
        _logger.info("Start update holiday ..............................")
        holi_lines = self.sudo().env["opsol_topnett.imp_conge"].search_read([], ['id', 'code_conge', 'matricule', 'date_start'])

        existing_holi = []
        new_holi = []
        for lp_ in transformed:
            if not lp_.get('matricule', False) or not lp_.get('date_start', False):
                continue

            exist = list(filter(
                lambda x: x['matricule'] and str(x['matricule']) == str(lp_['matricule']) and x['code_conge'] and str(x['code_conge']) == str(lp_['code_conge']) and x['date_start'] and str(x['date_start'])[:10] == str(lp_['date_start'])[:10], holi_lines))
            if exist:
                existing_holi.append((exist[0]['id'], lp_))
            else:
                new_holi.append(lp_)

        msg = f"{len(existing_holi)} absence(s) mise a jour et {len(new_holi)} nouvelles absences"

        # write ou update
        for _id, vals in existing_holi:
            self.env["opsol_topnett.imp_conge"].browse(_id).update(vals)

        # create new employee
        for elt in new_holi:
            self.env["opsol_topnett.imp_conge"].create(elt)

        _logger.info("end updating holiday..............................")
        return msg

    def transform_retraite(self, data_retraite, date_salaire=None):
        """Changement du nom des colonnes conges et adaptation des valeurs"""
        _logger.info("Start transform conge data ..............................")

        def v_get_value(value, key):
            if not value:                
                return value
            elif key in ['date_start', 'date_end']:                
                return get_date_formated(value)
            elif key == 'matricule':
                vals = value.split("/") if value and type(value) == str else []
                return len(vals) > 1 and vals[-1].strip() or type(value) == str and value.strip() or value
            else:
                return value.strip() if type(value) == str else value

        _logger.debug("Start transform..............................")
        map_croisement = {'matricule': 'matricule', 'EWC1   RETRAITE COMPL TR A': 'retra_comp_a', 'EWC2   RETRAITE COMPL TR B': 'retra_comp_b', 'date_salaire': 'date_salaire'}
        transformed = []
        for line in data_retraite:
            line['date_salaire'] = date_salaire or get_date_formated(datetime.today())
            transformed.append({ map_croisement.get(key): v_get_value(line[key], map_croisement.get(key)) for key in line.keys() if key in map_croisement.keys()})
        return transformed

    def update_retraite(self, transformed):
        """Recupere les retraites info existants et les met a jour."""
        _logger.info("Start update holiday ..............................")
        retraite_lines = self.sudo().env["opsol_topnett.retraite_info"].search_read([], ['id', 'matricule', 'date_salaire'])

        existing_retraite = []
        new_retraite = []
        for lp_ in transformed:
            if not lp_.get('matricule', False) or not lp_.get('date_salaire', False):
                continue

            exist = list(filter(lambda x: str(x.get('matricule', "")) == str(lp_['matricule']) and str(x.get('date_salaire', "********"))[:7] == str(lp_['date_salaire'])[:7], retraite_lines))
            if exist:
                existing_retraite.append((exist[0]['id'], lp_))
            else:
                new_retraite.append(lp_)

        msg = f"{len(existing_retraite)} info(s) retraite mise a jour et {len(new_retraite)} nouvelle(s) info(s) retraites"

        # write ou update
        for _id, vals in existing_retraite:
            self.env["opsol_topnett.retraite_info"].browse(_id).update(vals)

        # create new employee
        for elt in new_retraite:
            self.env["opsol_topnett.retraite_info"].create(elt)

        _logger.info("end updating retraite..............................")
        return msg

    def transform(self, datas, date_salaire=None):
        """Change les nom des colonnes de salaire."""
        def v_get_value(value, key):
            if key in ['standard_price', 'supplier_stock']:
                return float(value.replace(',', '.')) if value else 0
            elif key in ['image_1920']:
                if 'http://' not in value:
                    value = value and 'http://' + value or value
                return value
            else:
                return value

        _logger.info("Start transform..............................")
        map_croisement = {
            'Matricule': 'matricule', 'Salaire brut': 'salaire_brut', 'Travaillées': 'h_travailles', 'Complémentaires': 'h_complementaires',
            'Supplémentaires': 'h_supplementaires', 'Maladie': 'c_maladie', 'A.T.': 'c_at', 'Maternité': 'c_maternite', 'Congés.P': 'c_payes', 'Autres Abs.': 'c_autres',
            'DateEntree1': 'date_entree1', 'DateEntree2': 'date_entree2', 'Maintien': 'maintien', 'Prime': 'prime', 'date_salaire': 'date_salaire'
        }
        transformed = []
        for line in datas:
            line['date_salaire'] = date_salaire or line.get('date_salaire', None) or date.today().strftime(DATE_FORMAT)
            transformed.append({ map_croisement.get(key): v_get_value(line[key], map_croisement.get(key)) for key in line.keys() if key in map_croisement.keys()})

        _logger.info(f"{str(transformed[:5])}")
        _logger.info("end transform..............................")
        return transformed

    def transform2(self, lines):
        """Adapte les valeurs des salaires."""

        _logger.info("Start transform2..............................")
        transformed = []
        for line in lines:
            vals = line['matricule'].split("/") if type(line['matricule']) == str else []
            line['matricule'] = len(vals) > 1 and vals[-1].strip() or line['matricule']
            line['salaire_brut'] = float(line['salaire_brut'].replace(";", "")) if type(line['salaire_brut']) == 'str' else line['salaire_brut']
            line['date_entree1'] = get_date_formated(line['date_entree1'])
            line['date_entree2'] = get_date_formated(line['date_entree2'])
            line['date_salaire'] = get_date_formated(line['date_salaire']) if 'date_salaire' in line.keys() else get_date_formated(datetime.today())
            transformed.append(line)

            if not line["date_entree1"]:
                line.pop('date_entree1')
            if not line["date_entree2"]:
                line.pop('date_entree2')

        _logger.info(f"{str(transformed[:5])}")
        _logger.info("end transform2..............................")
        return transformed

    def split_nouveau_existant(self, transformed):
        """Separe creation et mise a jour."""
        _logger.info(f"data : {transformed[:5]}")
        _logger.info("Start Separation nouveau et anciens ..............................")

        imp_lines = self.env["opsol_topnett.imp_salaire_line"].search_read([('date_salaire', '!=', False)], ['id', 'date_salaire', 'matricule'])
        existing_products = []
        new_products = []
        for lp_ in transformed:
            if 'date_salaire' not in lp_.keys() or type(lp_['date_salaire']) == float:
                continue

            exist = list(filter(lambda x: x['date_salaire'] and lp_['date_salaire'] and x['matricule'] == lp_['matricule'] and (type(x['date_salaire']) == str and  x['date_salaire'][:7] or x['date_salaire'].strftime('%Y-%m')) == lp_['date_salaire'][:7], imp_lines))
            if exist:
                existing_products.append((exist[0]['id'], lp_))
            else:
                new_products.append(lp_)
        return {'exist': existing_products, 'news': new_products}

    def creer_line(self, imp_create):
        """Creation des nouvelles lignes de salaire."""

        _logger.info("Start Creer nouvelles lignes ..............................")
        elts = self.env["opsol_topnett.imp_salaire_line"]
        for line in imp_create:
            elts |= self.env["opsol_topnett.imp_salaire_line"].create(line)
        elts.with_context(skip_error=True).generate_payslip()
        _logger.info("compute lot....")
        lots = elts.mapped('lot_id')
        lots.action_reload_and_recompute_payslip()
        _logger.info("end of pipeline!")

    def mettre_a_jour_line(self, update_salaire):
        """Mise a jour des importation salaire."""
        _logger.info("Start update existing products..............................")
        ids = []
        for _id, vals in update_salaire:
            self.env["opsol_topnett.imp_salaire_line"].browse(_id).update(vals)
            ids.append(_id)
        records = self.env["opsol_topnett.imp_salaire_line"].browse(ids)
        records.mapped('bulletin_id').action_payslip_cancel()
        records.mapped('bulletin_id').unlink()
        records.with_context(skip_error=True).generate_payslip()
        _logger.info("end of pipeline!")

    def execute_load_employee(self, document):  
        # load employee
        url = document.raw
        datas, date_salaire = self.file_to_dict(url, handler=document.handler, filename=document.name)
        transformed = self.transform_employee(datas)
        msg = self.update_employee(transformed)
        return msg

    def execute_load_holiday(self, document):
        # load holiday
        url = document.raw
        datas, date_salaire = self.file_to_dict(url, handler=document.handler, filename=document.name)
        transformed = self.transform_holiday(datas)
        msg = self.update_holiday(transformed)
        return msg

    def execute_load_retraite(self, document):
        # load retaite
        url = document.raw
        datas, date_salaire = self.file_to_dict(
            url, line_header=3, line_first_row=2, handler="openxl",
            column_header=['matricule', 'retra_comp_a', 'retra_comp_b'], column_index=[0,40,41],
            filename=document.name
        )
        transformed = self.transform_retraite(datas, date_salaire=date_salaire)
        msg = self.update_retraite(transformed)
        return msg

    def execute_load_salaire(self, document):
        # load salaire
        url = document.raw
        datas, date_salaire = self.file_to_dict(url, handler=document.handler, filename=document.name)
        transformed = self.transform(datas, date_salaire=date_salaire)
        transformed = self.transform2(transformed)
        result = self.split_nouveau_existant(transformed)
        self.mettre_a_jour_line(result['exist'])
        self.creer_line(result['news'])
        msg = f"{len(result['exist'])} Importations mises a jour et {len(result['news'])} Nouvelles Importations"
        return msg
